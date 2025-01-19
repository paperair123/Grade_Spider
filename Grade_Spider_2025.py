import openpyxl
import requests
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook

# 定义URL和文件路径
url = "https://www.44886.com/BKOfficeMsgOut-12592.htm"
excel_path = "E:\\OneDrive\\桌面\\4班10A二阶成绩.xls"
output_excel_path = "E:\\OneDrive\\桌面\\成绩表.xlsx"

# 读取Excel文件
try:
    df = pd.read_excel(excel_path, header=0)
    print("开始读取二要素，共" + str(len(df)) + "条数据")
except FileNotFoundError:
    print(f"文件未找到: {excel_path}")
    exit(1)
except Exception as e:
    print(f"读取Excel文件时发生错误: {e}")
    exit(1)

# 加载目标Excel文件
try:
    wb = load_workbook(output_excel_path)
    ws = wb.active
except FileNotFoundError:
    print(f"输出文件未找到: {output_excel_path}, 创建新文件")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
except Exception as e:
    print(f"加载Excel文件时发生错误: {e}")
    exit(1)

# 设置HTTP请求头
headers = {
    'User-Agent': "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/131.0.0.0 Safari/537.36 Edg/131.0.0.0",
    'Accept': "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.7",
    'Accept-Encoding': "gzip, deflate, br, zstd",
    'Content-Type': "application/x-www-form-urlencoded",
    'cache-control': "max-age=0",
    'sec-ch-ua': "\"Microsoft Edge\";v=\"131\", \"Chromium\";v=\"131\", \"Not_A Brand\";v=\"24\"",
    'sec-ch-ua-mobile': "?0",
    'sec-ch-ua-platform': "\"Windows\"",
    'origin': "https://www.44886.com",
    'upgrade-insecure-requests': "1",
    'sec-fetch-site': "same-origin",
    'sec-fetch-mode': "navigate",
    'sec-fetch-user': "?1",
    'sec-fetch-dest': "document",
    'referer': "https://www.44886.com/BKOfficeMsgOut-12592.htm?nothing",
    'accept-language': "zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6",
    'priority': "u=0, i",
    'Cookie': "bbs_sid=hvm3cihle4h0ak39e37vuj0s8m; cookie_test=OKlDrdD7gC1XOA51acMETekwV2cv4BVIZ94MB0UJdkRh3Vrw"
}

# 遍历DataFrame中的每一行
for index, ser in df.iterrows():
    name = ser.iloc[0]  # 第一列的值
    examid = ser.iloc[1]  # 第二列的值

    # 构建payload
    pre_payload = f"c0={name}&c2={examid}"
    payload = pre_payload.encode('utf-8')

    try:
        # 发送POST请求
        response = requests.post(url, data=payload, headers=headers)
        response.raise_for_status()  # 检查请求是否成功

        # 解析HTML内容
        soup = BeautifulSoup(response.text, 'html.parser')

        # 提取col-4和col-8元素
        keys = soup.find_all(class_='col-4')
        values = soup.find_all(class_='col-8')

        if len(keys) != len(values):
            print(f"警告: 数据不匹配，跳过第{index + 1}行")
            continue

        # 将键值对组合成字典
        data_dict = dict(zip([k.get_text(strip=True) for k in keys], [v.get_text(strip=True) for v in values]))

        # 将数据添加到工作表
        row_data = []
        for key in keys:
            key_text = key.get_text(strip=True)
            row_data.append(key_text)
            row_data.append(data_dict[key_text])

        ws.append(row_data)
        print(f"成功处理第{index + 1}行")

    except requests.exceptions.RequestException as req_err:
        print(f"请求错误: {req_err}")
    except Exception as e:
        print(f"处理第{index + 1}行时发生错误: {e}")

# 保存最终结果
wb.save(output_excel_path)
print("完成")



